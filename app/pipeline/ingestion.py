import fitz  # PyMuPDF
from docx import Document as DocxDocument
from bs4 import BeautifulSoup


def parse_pdf(file_path: str) -> str:
    """Extract all text from a PDF file."""
    doc = fitz.open(file_path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text


def parse_docx(file_path: str) -> str:
    """Extract all text from a DOCX file."""
    doc = DocxDocument(file_path)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text


def parse_html(file_path: str) -> str:
    """Extract all text from an HTML file."""
    with open(file_path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")
    return soup.get_text(separator="\n", strip=True)


def parse_txt(file_path: str) -> str:
    """Extract all text from a TXT file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def parse_xlsx(file_path: str) -> str:
    """Extract all text from an XLSX file (reads all sheets, row by row)."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True)
    text = ""
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        text += f"\n--- Sheet: {sheet} ---\n"
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text += row_text + "\n"
    wb.close()
    return text


def parse_document(file_path: str, file_type: str) -> str:
    """Parse document based on file type. Returns extracted text."""
    if file_type == "pdf":
        return parse_pdf(file_path)
    elif file_type == "docx":
        return parse_docx(file_path)
    elif file_type == "html":
        return parse_html(file_path)
    elif file_type == "txt":
        return parse_txt(file_path)
    elif file_type == "xlsx":
        return parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")